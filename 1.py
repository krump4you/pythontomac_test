import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import sys
import green
import os
import csv
from openpyxl import load_workbook, Workbook
from pathlib import Path
from tkinter import ttk


def redirect_stdout(text_widget):
    class StdoutRedirector:
        def __init__(self, widget):
            self.widget = widget

        def write(self, message):
            self.widget.insert("end", message)
            self.widget.see("end")

        def flush(self):
            pass

    sys.stdout = StdoutRedirector(text_widget)


file_entry = None
progress_bar = None
processing_label = None


def browse_file():
    global file_entry
    file_path = filedialog.askopenfilename(filetypes=(
        ("Excel files", "*.xlsx"), ("CSV files", "*.csv")))
    file_entry.delete(0, tk.END)
    file_entry.insert(tk.END, file_path)


def split_file():
    print("Начинаем...")
    global file_entry
    global progress_bar
    global processing_label

    input_file = file_entry.get()

    if not input_file:
        messagebox.showerror("Error", "No file selected.")
        return

    path = Path(input_file)
    ishod = path.stem
    _, file_extension = os.path.splitext(input_file)

    if file_extension == ".xlsx":
        workbook = load_workbook(filename=input_file)
        sheet = workbook.active
        totalr = sheet.max_row

        rows = sheet.iter_rows(values_only=True)
    elif file_extension == ".csv":
        rows = csv.reader(open(input_file, "r"))
    else:
        messagebox.showerror(
            "Error", "Invalid file format. Please provide a valid XLSX or CSV file.")
        return

    valid_rows = []
    invalid_rows = []
    for k, row in enumerate(rows):
        def long_running_task():
            progress = (k + 1) / totalr * 100
            progress_bar['value'] = progress

            if progress % 2 == 0:
                if progress % 4 == 0:
                    processing_label.config(text="")
                else:
                    processing_label.config(text="Processing")

            if progress % 10 == 0:
                print(f"Фильтруем..., прогресс:{progress:.0f}%")

            # Update the GUI
            root.update()

        long_running_task()

        phone_number = row[0]
        is_valid = green.checkWA(phone_number)
        if is_valid == "skip":
            print(f"Ошибка проверки номера: {phone_number}...Пропускаем...", row)
            continue

        if is_valid:
            valid_rows.append(row)
        else:
            invalid_rows.append(row)

    valid_output_file = ishod + "_BOTC" + file_extension
    invalid_output_file = ishod + "_NO_BOTC" + file_extension

    if file_extension == ".xlsx":
        workbook_valid = Workbook()
        sheet_valid = workbook_valid.active
        for row in valid_rows:
            sheet_valid.append(row)
        workbook_valid.save(valid_output_file)

        workbook_invalid = Workbook()
        sheet_invalid = workbook_invalid.active
        for row in invalid_rows:
            sheet_invalid.append(row)
        workbook_invalid.save(invalid_output_file)

    elif file_extension == ".csv":
        with open(valid_output_file, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(valid_rows)

        with open(invalid_output_file, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(invalid_rows)

    messagebox.showinfo(
        "Фильтрация окончена!", f"Отфильтрованные базы находятся в одной папке с фильтром.\nВОТСЫ: {valid_output_file}\nБез ВОТСов: {invalid_output_file}")

    # Reset the progress bar after the operation is complete
    progress_bar['value'] = 0
    processing_label.config(text="")


def main():
    global file_entry
    global progress_bar
    global processing_label

    root = tk.Tk()
    root.title("Фильтр ВОТС")

    file_label = tk.Label(root, text="Выберите базу:")
    file_label.pack()

    file_entry = tk.Entry(root, width=30)
    file_entry.pack(pady=5)

    browse_button = tk.Button(root, text="Загрузить", command=browse_file)
    browse_button.pack(pady=5)

    split_button = tk.Button(root, text="Фильтровать", command=split_file)
    split_button.pack(pady=5)

    progress_bar = ttk.Progressbar(
        root,
        orient='horizontal',
        mode='determinate',
        length=280
    )
    progress_bar.pack(pady=5)

    processing_label = tk.Label(root, text="")
    processing_label.pack(pady=5)

    text_widget = tk.Text(root, height=2)
    text_widget.pack(fill='both', expand=1, pady=5)

    redirect_stdout(text_widget)

    exit_button = tk.Button(root, text="Выход", command=root.quit)
    exit_button.pack(pady=5)

    root.mainloop()
    exit(0)


if __name__ == "__main__":
    main()